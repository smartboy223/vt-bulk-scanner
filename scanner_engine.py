"""
VirusTotal Bulk Scanner Engine
Multi-key parallel VT scanning with cache, quota management, retry, and resume.
"""

import os
import re
import json
import time
import uuid
import shutil
import threading
import queue
import requests
import logging
from datetime import datetime, date
from collections import deque

logger = logging.getLogger(__name__)

VT_FILE_URL = "https://www.virustotal.com/api/v3/files/{}"
VT_IP_URL = "https://www.virustotal.com/api/v3/ip_addresses/{}"
VT_DOMAIN_URL = "https://www.virustotal.com/api/v3/domains/{}"

FREE_REQUESTS_PER_MINUTE = 4
FREE_REQUESTS_PER_DAY = 500
CACHE_TTL_HOURS = 24
MAX_RETRIES = 3
RETRY_BACKOFF = [2, 5, 10]

HASH_MD5 = re.compile(r'^[a-fA-F0-9]{32}$')
HASH_SHA1 = re.compile(r'^[a-fA-F0-9]{40}$')
HASH_SHA256 = re.compile(r'^[a-fA-F0-9]{64}$')
IP_PATTERN = re.compile(
    r'^(?:(?:25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(?:25[0-5]|2[0-4]\d|[01]?\d\d?)$'
)
DOMAIN_PATTERN = re.compile(
    r'^(?:[a-zA-Z0-9](?:[a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}$'
)


def detect_indicator_type(value):
    value = value.strip()
    if HASH_SHA256.match(value):
        return 'sha256'
    if HASH_SHA1.match(value):
        return 'sha1'
    if HASH_MD5.match(value):
        return 'md5'
    if IP_PATTERN.match(value):
        return 'ip'
    if DOMAIN_PATTERN.match(value):
        return 'domain'
    return None


def defang(text):
    text = text.replace('hxxp://', '').replace('hxxps://', '')
    text = text.replace('http://', '').replace('https://', '')
    text = text.replace('[.]', '.').replace('[dot]', '.')
    text = text.replace('(.)', '.')
    text = re.sub(r'\s+\.\s+', '.', text)
    text = text.split('/')[0]
    text = text.strip(' \t\r\n"\'`<>[](){}')
    return text


def parse_indicators(text):
    result = parse_indicators_detailed(text)
    return result['indicators']


def parse_indicators_detailed(text):
    indicators = []
    seen = set()
    rejected = []
    duplicates = []
    auto_fixed = []

    line_num = 0
    for line in text.strip().splitlines():
        line_num += 1
        raw = line
        line = line.strip()
        if not line or line.startswith('#'):
            continue

        hash_match = re.search(
            r'\b([a-fA-F0-9]{64}|[a-fA-F0-9]{40}|[a-fA-F0-9]{32})\b', line
        )
        if hash_match:
            value = hash_match.group(1).lower()
            ioc_type = detect_indicator_type(value)
            if ioc_type:
                if value in seen:
                    duplicates.append({
                        'line': line_num, 'text': raw.strip(),
                        'value': value, 'reason': 'Duplicate',
                    })
                    continue
                comment = line.replace(hash_match.group(1), '').strip(' ,;:\t')
                seen.add(value)
                indicators.append({
                    'value': value, 'type': ioc_type,
                    'comment': comment, 'status': 'pending',
                    'result': None, 'from_cache': False,
                })
                continue

        cleaned = defang(line)

        tokens = re.split(r'[\s,;|]+', cleaned)
        found = False
        for token in tokens:
            token = token.strip().strip('"\'`<>[](){}')
            if not token:
                continue
            ioc_type = detect_indicator_type(token)
            if ioc_type:
                val = token.lower() if ioc_type in ('md5', 'sha1', 'sha256') else token
                if val in seen:
                    duplicates.append({
                        'line': line_num, 'text': raw.strip(),
                        'value': val, 'reason': 'Duplicate',
                    })
                else:
                    seen.add(val)
                    was_fixed = cleaned != line.strip()
                    indicators.append({
                        'value': token, 'type': ioc_type,
                        'comment': '', 'status': 'pending',
                        'result': None, 'from_cache': False,
                    })
                    if was_fixed:
                        auto_fixed.append({
                            'line': line_num,
                            'original': raw.strip(),
                            'fixed': token,
                        })
                found = True
                break

        if not found:
            reason = 'No valid indicator found'
            stripped = line.strip()
            if len(stripped) < 4:
                reason = 'Too short'
            elif len(stripped) > 256:
                reason = 'Too long'
            rejected.append({
                'line': line_num, 'text': stripped[:80], 'reason': reason,
            })

    total_lines = line_num
    return {
        'indicators': indicators,
        'rejected': rejected,
        'duplicates': duplicates,
        'auto_fixed': auto_fixed,
        'total_lines': total_lines,
    }


# ---------------------------------------------------------------------------
# VT Cache
# ---------------------------------------------------------------------------

class VTCache:
    def __init__(self, cache_file, ttl_hours=CACHE_TTL_HOURS):
        self.cache_file = cache_file
        self.ttl = ttl_hours * 3600
        self.lock = threading.Lock()
        self._load()

    def _load(self):
        if os.path.exists(self.cache_file):
            try:
                with open(self.cache_file, 'r', encoding='utf-8') as f:
                    self.data = json.load(f)
            except Exception:
                self.data = {}
        else:
            self.data = {}

    def _save(self):
        tmp = self.cache_file + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(self.data, f)
        if os.path.exists(self.cache_file):
            os.replace(tmp, self.cache_file)
        else:
            os.rename(tmp, self.cache_file)

    def get(self, indicator):
        with self.lock:
            key = indicator.lower().strip()
            entry = self.data.get(key)
            if entry and time.time() - entry.get('cached_at', 0) < self.ttl:
                return entry.get('result')
            return None

    def set(self, indicator, result):
        with self.lock:
            key = indicator.lower().strip()
            self.data[key] = {
                'result': result,
                'cached_at': time.time(),
                'cached_date': datetime.now().isoformat(),
            }
            self._save()

    def has(self, indicator):
        return self.get(indicator) is not None

    def count_valid(self):
        with self.lock:
            now = time.time()
            return sum(
                1 for v in self.data.values()
                if now - v.get('cached_at', 0) < self.ttl
            )

    def count_hits(self, indicators):
        hits = 0
        for ind in indicators:
            if self.has(ind['value']):
                hits += 1
        return hits

    def clear_all(self):
        with self.lock:
            count = len(self.data)
            self.data = {}
            self._save()
        return count

    def clear_expired(self):
        with self.lock:
            now = time.time()
            expired = [
                k for k, v in self.data.items()
                if now - v.get('cached_at', 0) >= self.ttl
            ]
            for k in expired:
                del self.data[k]
            if expired:
                self._save()
            return len(expired)

    def size_bytes(self):
        try:
            if os.path.exists(self.cache_file):
                return os.path.getsize(self.cache_file)
        except OSError:
            pass
        return 0


# ---------------------------------------------------------------------------
# Per-Key Rate Limiter
# ---------------------------------------------------------------------------

class KeyRateLimiter:
    def __init__(self):
        self.minute_ts = deque()
        self.daily_count = 0
        self.daily_date = date.today()
        self.exhausted = False
        self.rl_until = 0

    def _reset_daily(self):
        if date.today() > self.daily_date:
            self.daily_count = 0
            self.daily_date = date.today()
            self.exhausted = False

    def _clean_window(self):
        now = time.time()
        while self.minute_ts and now - self.minute_ts[0] >= 60:
            self.minute_ts.popleft()

    def can_request(self):
        self._reset_daily()
        if self.exhausted or self.daily_count >= FREE_REQUESTS_PER_DAY:
            self.exhausted = True
            return False
        if self.rl_until > time.time():
            return False
        self._clean_window()
        return len(self.minute_ts) < FREE_REQUESTS_PER_MINUTE

    def wait_time(self):
        self._reset_daily()
        if self.exhausted:
            return -1
        now = time.time()
        if self.rl_until > now:
            return self.rl_until - now
        self._clean_window()
        if len(self.minute_ts) >= FREE_REQUESTS_PER_MINUTE:
            return 60 - (now - self.minute_ts[0]) + 1
        return 0

    def record_use(self):
        self.minute_ts.append(time.time())
        self.daily_count += 1

    def mark_rate_limited(self):
        self.rl_until = time.time() + 62

    def is_daily_exhausted(self):
        self._reset_daily()
        return self.exhausted or self.daily_count >= FREE_REQUESTS_PER_DAY


# ---------------------------------------------------------------------------
# VT Quota Check
# ---------------------------------------------------------------------------

VT_USER_URL = "https://www.virustotal.com/api/v3/users/{}"


def check_api_quota(api_key):
    try:
        headers = {"x-apikey": api_key}
        r = requests.get(VT_USER_URL.format(api_key), headers=headers, timeout=15)
        if not r.ok:
            return None
        data = r.json()
        quotas = data.get('data', {}).get('attributes', {}).get('quotas', {})
        daily = quotas.get('api_requests_daily', {})
        monthly = quotas.get('api_requests_monthly', {})
        return {
            'daily_used': daily.get('used', 0),
            'daily_allowed': daily.get('allowed', 0),
            'daily_remaining': max(0, daily.get('allowed', 0) - daily.get('used', 0)),
            'monthly_used': monthly.get('used', 0),
            'monthly_allowed': monthly.get('allowed', 0),
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Scan Job
# ---------------------------------------------------------------------------

class ScanJob:
    def __init__(self, scan_id, name, indicators, scan_dir, force_rescan=False):
        self.scan_id = scan_id
        self.name = name
        self.indicators = indicators
        self.scan_dir = scan_dir
        self.force_rescan = force_rescan
        self.status = 'created'
        self.pause_reason = ''
        self.created_at = datetime.now().isoformat()
        self.updated_at = datetime.now().isoformat()
        self.current_index = 0
        self.error_message = ''
        self.log = []
        self.started_at = None
        self.last_activity_at = None
        self.wait_message = ''
        self._done_at_start = 0
        self._recent_vt = deque()
        self._recent_cache = deque()
        self._WINDOW = 120
        os.makedirs(scan_dir, exist_ok=True)

    def record_completion(self, from_cache=False):
        now = time.time()
        self.last_activity_at = now
        if from_cache:
            self._recent_cache.append(now)
        else:
            self._recent_vt.append(now)

    def _prune_window(self):
        cutoff = time.time() - self._WINDOW
        while self._recent_vt and self._recent_vt[0] < cutoff:
            self._recent_vt.popleft()
        while self._recent_cache and self._recent_cache[0] < cutoff:
            self._recent_cache.popleft()

    @property
    def total(self):
        return len(self.indicators)

    @property
    def completed_count(self):
        return sum(1 for i in self.indicators if i['status'] == 'completed')

    @property
    def failed_count(self):
        return sum(1 for i in self.indicators if i['status'] == 'failed')

    @property
    def pending_count(self):
        return sum(1 for i in self.indicators if i['status'] == 'pending')

    @property
    def cached_count(self):
        return sum(1 for i in self.indicators if i.get('from_cache'))

    @property
    def progress_percent(self):
        if self.total == 0:
            return 100
        return round(
            (self.completed_count + self.failed_count) / self.total * 100, 1
        )

    @property
    def elapsed_seconds(self):
        if not self.started_at:
            return 0
        return time.time() - self.started_at

    @property
    def speed_per_minute(self):
        self._prune_window()
        recent_total = len(self._recent_vt) + len(self._recent_cache)
        if recent_total == 0:
            elapsed = self.elapsed_seconds
            if elapsed < 5:
                return 0
            done = self.completed_count + self.failed_count - self._done_at_start
            if done <= 0:
                return 0
            return round(done / (elapsed / 60), 1)
        window_span = min(self._WINDOW, self.elapsed_seconds)
        if window_span < 3:
            return 0
        return round(recent_total / (window_span / 60), 1)

    @property
    def eta_seconds(self):
        self._prune_window()
        pending = self.pending_count
        if pending == 0:
            return 0
        vt_count = len(self._recent_vt)
        cache_count = len(self._recent_cache)
        window_span = min(self._WINDOW, max(self.elapsed_seconds, 1))
        if vt_count + cache_count == 0:
            spd = self.speed_per_minute
            if spd <= 0:
                num_keys = max(1, len(self._recent_vt) or 1)
                return int(pending / (4 * num_keys) * 60)
            return int(pending / spd * 60)
        vt_rate = vt_count / window_span if vt_count else 0
        cache_rate = cache_count / window_span if cache_count else 0
        if vt_rate + cache_rate > 0:
            return int(pending / (vt_rate + cache_rate))
        return 0

    def save_state(self):
        state = {
            'scan_id': self.scan_id,
            'name': self.name,
            'status': self.status,
            'pause_reason': self.pause_reason,
            'force_rescan': self.force_rescan,
            'created_at': self.created_at,
            'updated_at': datetime.now().isoformat(),
            'current_index': self.current_index,
            'error_message': self.error_message,
            'indicators': self.indicators,
            'log': self.log[-200:],
        }
        state_file = os.path.join(self.scan_dir, 'state.json')
        tmp = state_file + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, default=str)
        if os.path.exists(state_file):
            os.replace(tmp, state_file)
        else:
            os.rename(tmp, state_file)

    @classmethod
    def load_state(cls, scan_dir):
        state_file = os.path.join(scan_dir, 'state.json')
        with open(state_file, 'r', encoding='utf-8') as f:
            state = json.load(f)
        job = cls(
            state['scan_id'], state['name'],
            state['indicators'], scan_dir,
            state.get('force_rescan', False),
        )
        job.status = state['status']
        job.pause_reason = state.get('pause_reason', '')
        job.created_at = state['created_at']
        job.updated_at = state.get('updated_at', '')
        job.current_index = state.get('current_index', 0)
        job.error_message = state.get('error_message', '')
        job.log = state.get('log', [])
        return job

    def add_log(self, message):
        self.log.append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'message': message,
        })
        if len(self.log) > 400:
            self.log = self.log[-200:]

    def to_summary(self):
        return {
            'scan_id': self.scan_id,
            'name': self.name,
            'status': self.status,
            'pause_reason': self.pause_reason,
            'created_at': self.created_at,
            'updated_at': self.updated_at,
            'total': self.total,
            'completed': self.completed_count,
            'failed': self.failed_count,
            'pending': self.pending_count,
            'cached': self.cached_count,
            'progress': self.progress_percent,
            'error_message': self.error_message,
            'elapsed_seconds': int(self.elapsed_seconds),
            'speed_per_minute': self.speed_per_minute,
            'eta_seconds': self.eta_seconds,
            'wait_message': self.wait_message,
        }


# ---------------------------------------------------------------------------
# Scan Engine
# ---------------------------------------------------------------------------

class ScanEngine:
    def __init__(self, data_dir='data'):
        self.data_dir = data_dir
        self.scans_dir = os.path.join(data_dir, 'scans')
        self.config_file = os.path.join(data_dir, 'config.json')
        os.makedirs(self.scans_dir, exist_ok=True)

        self.active_scans = {}
        self.scan_jobs = {}
        self.stop_flags = {}

        self._load_config()
        self.cache = VTCache(os.path.join(data_dir, 'vt_cache.json'))
        self._load_existing_scans()
        self._start_auto_resume_monitor()

    def _load_config(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r') as f:
                self.config = json.load(f)
        else:
            self.config = {'api_keys': []}
            self._save_config()

    def _save_config(self):
        os.makedirs(self.data_dir, exist_ok=True)
        with open(self.config_file, 'w') as f:
            json.dump(self.config, f, indent=2)

    def _load_existing_scans(self):
        if not os.path.exists(self.scans_dir):
            return
        for scan_id in os.listdir(self.scans_dir):
            scan_dir = os.path.join(self.scans_dir, scan_id)
            state_file = os.path.join(scan_dir, 'state.json')
            if os.path.isdir(scan_dir) and os.path.exists(state_file):
                try:
                    job = ScanJob.load_state(scan_dir)
                    if job.status == 'running':
                        job.status = 'paused'
                        job.pause_reason = 'interrupted'
                        job.add_log('Scan interrupted - marked paused for resume')
                        job.save_state()
                    self.scan_jobs[scan_id] = job
                except Exception as e:
                    logger.error("Failed to load scan %s: %s", scan_id, e)

    def _start_auto_resume_monitor(self):
        def monitor():
            while True:
                time.sleep(60)
                try:
                    self._load_config()
                    for sid, job in list(self.scan_jobs.items()):
                        if (
                            job.status == 'paused'
                            and job.pause_reason == 'quota_exhausted'
                            and self.config['api_keys']
                        ):
                            alive = (
                                sid in self.active_scans
                                and self.active_scans[sid].is_alive()
                            )
                            if not alive:
                                job.add_log(
                                    'Auto-resume: checking for available keys...'
                                )
                                self.start_scan(sid)
                except Exception:
                    pass

        t = threading.Thread(target=monitor, daemon=True)
        t.start()

    # -- API Key management --

    def get_api_keys(self):
        return self.config['api_keys']

    def add_api_key(self, key, label=''):
        key = key.strip()
        if not re.match(r'^[a-fA-F0-9]{64}$', key):
            return False, 'Invalid API key format (must be 64 hex characters)'
        for existing in self.config['api_keys']:
            if existing['key'] == key:
                return False, 'API key already exists'
        self.config['api_keys'].append({
            'key': key,
            'label': label or f'Key {len(self.config["api_keys"]) + 1}',
            'added_at': datetime.now().isoformat(),
        })
        self._save_config()
        return True, 'API key added successfully'

    def remove_api_key(self, key):
        self.config['api_keys'] = [
            k for k in self.config['api_keys'] if k['key'] != key
        ]
        self._save_config()
        return True, 'API key removed'

    # -- Cache --

    def get_cache_stats(self):
        size = self.cache.size_bytes()
        if size > 1048576:
            size_str = f'{size / 1048576:.1f} MB'
        elif size > 1024:
            size_str = f'{size / 1024:.1f} KB'
        else:
            size_str = f'{size} B'
        return {
            'valid': self.cache.count_valid(),
            'total': len(self.cache.data),
            'size': size_str,
        }

    def count_cache_hits(self, indicators):
        return self.cache.count_hits(indicators)

    def clear_cache(self):
        count = self.cache.clear_all()
        return True, f'Cache cleared ({count} entries removed)'

    def clear_expired_cache(self):
        count = self.cache.clear_expired()
        return True, f'{count} expired entries removed'

    def check_all_quotas(self):
        results = []
        for ki in self.config['api_keys']:
            q = check_api_quota(ki['key'])
            results.append({
                'label': ki.get('label', ki['key'][:8]),
                'key_short': ki['key'][:8] + '...' + ki['key'][-4:],
                'quota': q,
            })
        return results

    # -- Scan management --

    def create_scan(self, name, text_content, force_rescan=False):
        indicators = parse_indicators(text_content)
        if not indicators:
            return None, 'No valid indicators found in the input', {}

        cache_hits = self.cache.count_hits(indicators) if not force_rescan else 0

        scan_id = str(uuid.uuid4())[:8]
        scan_dir = os.path.join(self.scans_dir, scan_id)

        job = ScanJob(scan_id, name, indicators, scan_dir, force_rescan)
        job.save_state()
        self.scan_jobs[scan_id] = job

        stats = {
            'total': len(indicators),
            'cache_hits': cache_hits,
            'new_lookups': len(indicators) - cache_hits,
        }

        return scan_id, f'Scan created with {len(indicators)} indicators ({cache_hits} cached)', stats

    def get_scan(self, scan_id):
        return self.scan_jobs.get(scan_id)

    def get_all_scans(self):
        scans = []
        for _sid, job in sorted(
            self.scan_jobs.items(),
            key=lambda x: x[1].created_at,
            reverse=True,
        ):
            scans.append(job.to_summary())
        return scans

    def start_scan(self, scan_id):
        job = self.scan_jobs.get(scan_id)
        if not job:
            return False, 'Scan not found'
        if not self.config['api_keys']:
            return False, 'No API keys configured.'
        if scan_id in self.active_scans and self.active_scans[scan_id].is_alive():
            return False, 'Scan is already running'

        job.status = 'running'
        job.pause_reason = ''
        job.error_message = ''
        job.started_at = time.time()
        job._done_at_start = job.completed_count + job.failed_count
        job.save_state()

        stop_flag = threading.Event()
        self.stop_flags[scan_id] = stop_flag

        thread = threading.Thread(
            target=self._run_scan, args=(job, stop_flag), daemon=True
        )
        self.active_scans[scan_id] = thread
        thread.start()

        return True, 'Scan started'

    def pause_scan(self, scan_id):
        if scan_id in self.stop_flags:
            self.stop_flags[scan_id].set()
            return True, 'Pause requested'
        return False, 'Scan not running'

    def delete_scan(self, scan_id):
        if scan_id in self.stop_flags:
            self.stop_flags[scan_id].set()
        self.scan_jobs.pop(scan_id, None)
        self.active_scans.pop(scan_id, None)
        self.stop_flags.pop(scan_id, None)
        scan_dir = os.path.join(self.scans_dir, scan_id)
        if os.path.exists(scan_dir):
            shutil.rmtree(scan_dir)
        return True, 'Scan deleted'

    # -- Parallel scan runner --

    def _run_scan(self, job, stop_flag):
        api_keys = list(self.config['api_keys'])
        num_workers = len(api_keys)

        job.add_log(f'Scan started with {num_workers} API key(s) (parallel)')
        job.add_log(f'{job.pending_count} indicators remaining')
        if not job.force_rescan:
            job.add_log('Cache enabled - cached results will be reused')

        work_q = queue.Queue()
        for i, ind in enumerate(job.indicators):
            if ind['status'] not in ('completed', 'failed'):
                work_q.put(i)

        if work_q.empty():
            job.status = 'completed'
            job.add_log('All indicators already processed')
            job.save_state()
            return

        state_lock = threading.Lock()
        worker_exit_reasons = {}

        def worker(key_info, worker_id):
            api_key = key_info['key']
            key_label = key_info.get('label', f'Key {worker_id+1}')
            limiter = KeyRateLimiter()
            exit_reason = 'done'

            try:
                while not stop_flag.is_set():
                    try:
                        idx = work_q.get(timeout=1)
                    except queue.Empty:
                        break

                    indicator = job.indicators[idx]

                    if not job.force_rescan:
                        cached = self.cache.get(indicator['value'])
                        if cached:
                            with state_lock:
                                indicator['status'] = 'completed'
                                indicator['result'] = cached
                                indicator['from_cache'] = True
                                job.record_completion(from_cache=True)
                                job.add_log(
                                    f'[{job.completed_count}/{job.total}] '
                                    f'{indicator["type"].upper()} '
                                    f'{indicator["value"][:20]}... '
                                    f'(cached)'
                                )
                                job.save_state()
                            continue

                    processed = False
                    attempts = 0
                    while not processed and attempts < 40:
                        if stop_flag.is_set():
                            work_q.put(idx)
                            exit_reason = 'stopped'
                            worker_exit_reasons[worker_id] = exit_reason
                            return

                        if limiter.is_daily_exhausted():
                            with state_lock:
                                job.add_log(
                                    f'[W{worker_id+1}] Key {key_label} daily quota exhausted'
                                )
                            work_q.put(idx)
                            exit_reason = 'quota_exhausted'
                            worker_exit_reasons[worker_id] = exit_reason
                            return

                        wt = limiter.wait_time()
                        if wt > 0:
                            with state_lock:
                                job.wait_message = (
                                    f'Worker {worker_id+1} ({key_label}): '
                                    f'waiting {int(wt)}s for rate limit...'
                                )
                            for _ in range(int(wt) + 1):
                                if stop_flag.is_set():
                                    work_q.put(idx)
                                    exit_reason = 'stopped'
                                    worker_exit_reasons[worker_id] = exit_reason
                                    return
                                time.sleep(1)
                            attempts += 1
                            continue

                        if not limiter.can_request():
                            attempts += 1
                            time.sleep(1)
                            continue

                        with state_lock:
                            job.wait_message = (
                                f'Worker {worker_id+1}: querying '
                                f'{indicator["value"][:24]}...'
                            )

                        result = self._query_vt_with_retry(
                            indicator['value'], indicator['type'], api_key
                        )

                        if result['status_code'] == 429:
                            limiter.mark_rate_limited()
                            with state_lock:
                                job.add_log(
                                    f'[W{worker_id+1}] 429 on {key_label}, '
                                    f'backing off 62s'
                                )
                            attempts += 1
                            continue

                        limiter.record_use()

                        with state_lock:
                            if result['success']:
                                indicator['status'] = 'completed'
                                indicator['result'] = result['data']
                                self.cache.set(indicator['value'], result['data'])
                                r = result['data']
                                job.add_log(
                                    f'[{job.completed_count}/{job.total}] '
                                    f'{indicator["type"].upper()} '
                                    f'{indicator["value"][:20]}... '
                                    f'{r.get("rating","?")} '
                                    f'({r.get("detection_display", r.get("detection_ratio","?"))})'
                                )
                            else:
                                indicator['status'] = 'failed'
                                indicator['result'] = {
                                    'error': result.get('error', 'Unknown')
                                }
                                job.add_log(
                                    f'[{job.completed_count+job.failed_count}/{job.total}] '
                                    f'{indicator["value"][:20]}... '
                                    f'FAILED: {result.get("error","")}'
                                )
                            job.record_completion(from_cache=False)
                            job.wait_message = ''
                            job.save_state()

                        processed = True

                    if not processed:
                        with state_lock:
                            indicator['status'] = 'failed'
                            indicator['result'] = {'error': 'Max wait attempts'}
                            job.save_state()

            except Exception as e:
                exit_reason = f'error: {e}'
                with state_lock:
                    job.add_log(f'[W{worker_id+1}] Worker crashed: {e}')

            worker_exit_reasons[worker_id] = exit_reason

        try:
            threads = []
            for i, ki in enumerate(api_keys):
                t = threading.Thread(
                    target=worker, args=(ki, i), daemon=True
                )
                threads.append(t)
                t.start()

            for t in threads:
                t.join()

            if stop_flag.is_set():
                job.status = 'paused'
                job.pause_reason = 'user'
                job.add_log('Scan paused by user')
            elif job.pending_count == 0:
                job.status = 'completed'
                job.add_log(
                    f'Scan finished: {job.completed_count} OK, '
                    f'{job.failed_count} failed, '
                    f'{job.cached_count} from cache'
                )
            else:
                reasons = set(worker_exit_reasons.values())
                has_errors = any(r.startswith('error:') for r in reasons)
                has_quota = 'quota_exhausted' in reasons

                if has_quota:
                    job.status = 'paused'
                    job.pause_reason = 'quota_exhausted'
                    job.add_log(
                        'All keys exhausted daily quota. '
                        'Add more keys or resume tomorrow.'
                    )
                elif has_errors:
                    job.status = 'paused'
                    job.pause_reason = 'worker_error'
                    err_msgs = [r for r in reasons if r.startswith('error:')]
                    job.add_log(
                        f'Workers stopped due to errors: {"; ".join(err_msgs)}. '
                        f'Click Resume to retry.'
                    )
                else:
                    job.status = 'paused'
                    job.pause_reason = 'unknown'
                    job.add_log(
                        f'Workers exited with {job.pending_count} pending. '
                        f'Reasons: {reasons}. Click Resume to retry.'
                    )
            job.wait_message = ''
            job.save_state()

        except Exception as e:
            job.status = 'error'
            job.error_message = str(e)
            job.add_log(f'Error: {e}')
            job.save_state()
            logger.error("Scan %s failed: %s", job.scan_id, e, exc_info=True)

    # -- VT Query with retry --

    def _query_vt_with_retry(self, value, indicator_type, api_key):
        last_err = None
        for attempt in range(MAX_RETRIES):
            result = self._query_vt(value, indicator_type, api_key)
            if result['success'] or result['status_code'] == 429:
                return result
            if result['status_code'] == 404:
                return result
            last_err = result.get('error', 'Unknown')
            if attempt < MAX_RETRIES - 1:
                backoff = RETRY_BACKOFF[min(attempt, len(RETRY_BACKOFF) - 1)]
                time.sleep(backoff)
        return {
            'success': False, 'status_code': 0,
            'error': f'Failed after {MAX_RETRIES} retries: {last_err}',
        }

    def _query_vt(self, value, indicator_type, api_key):
        headers = {"x-apikey": api_key}
        try:
            if indicator_type in ('md5', 'sha1', 'sha256'):
                url = VT_FILE_URL.format(value)
            elif indicator_type == 'ip':
                url = VT_IP_URL.format(value)
            elif indicator_type == 'domain':
                url = VT_DOMAIN_URL.format(value)
            else:
                return {
                    'success': False, 'status_code': 0,
                    'error': f'Unknown type: {indicator_type}',
                }
            response = requests.get(url, headers=headers, timeout=30)

            if response.status_code == 429:
                return {'success': False, 'status_code': 429, 'error': 'Rate limited'}
            if response.status_code == 404:
                return {
                    'success': True, 'status_code': 404,
                    'data': {
                        'found': False, 'indicator': value,
                        'type': indicator_type, 'rating': 'unknown',
                        'positives': 0, 'total': 0,
                        'detection_ratio': '0/0',
                        'detection_display': 'Not found',
                        'malicious': 0, 'suspicious': 0,
                        'message': 'Not found in VirusTotal',
                    },
                }
            if not response.ok:
                return {
                    'success': False, 'status_code': response.status_code,
                    'error': f'HTTP {response.status_code}',
                }
            data = response.json()

            if indicator_type in ('md5', 'sha1', 'sha256'):
                parsed = self._parse_file_result(data, value)
            elif indicator_type == 'ip':
                parsed = self._parse_ip_result(data, value)
            elif indicator_type == 'domain':
                parsed = self._parse_domain_result(data, value)
            else:
                parsed = {'found': False}

            return {'success': True, 'status_code': 200, 'data': parsed}

        except requests.exceptions.Timeout:
            return {'success': False, 'status_code': 0, 'error': 'Timeout'}
        except requests.exceptions.ConnectionError:
            return {'success': False, 'status_code': 0, 'error': 'Connection error'}
        except Exception as e:
            return {'success': False, 'status_code': 0, 'error': str(e)}

    # -- Result parsers --

    def _parse_file_result(self, data, original_hash):
        result = {'found': True, 'type': 'file', 'indicator': original_hash}
        try:
            attrs = data.get('data', {}).get('attributes', {})
            stats = attrs.get('last_analysis_stats', {})

            malicious = stats.get('malicious', 0)
            suspicious = stats.get('suspicious', 0)
            undetected = stats.get('undetected', 0)
            harmless = stats.get('harmless', 0)
            total = malicious + suspicious + undetected + harmless

            names = list(attrs.get('names', []))
            meaningful = attrs.get('meaningful_name', '')
            if meaningful and meaningful in names:
                names.remove(meaningful)
                names.insert(0, meaningful)

            # Exiftool fields
            exif = attrs.get('exiftool', {})
            copyright_info = exif.get('LegalCopyright', '-')
            description = exif.get('FileDescription', '-')
            original_name = exif.get('OriginalFileName', exif.get('InternalName', '-'))
            product_name = exif.get('ProductName', '-')

            result.update({
                'md5': attrs.get('md5', '-'),
                'sha1': attrs.get('sha1', '-'),
                'sha256': attrs.get('sha256', '-'),
                'file_type': attrs.get('type_description', '-'),
                'file_size': attrs.get('size', 0),
                'file_names': ', '.join(names[:5]),
                'meaningful_name': meaningful or '-',
                'positives': malicious + suspicious,
                'total': total,
                'malicious': malicious,
                'suspicious': suspicious,
                'undetected': undetected,
                'harmless_count': harmless,
                'detection_ratio': f'{malicious + suspicious}/{total}' if total else '0/0',
                'detection_display': f'{malicious + suspicious}/{total}' if total else '0/0',
                'first_submission': self._ts(attrs.get('first_submission_date')),
                'last_submission': self._ts(attrs.get('last_submission_date')),
                'last_analysis': self._ts(attrs.get('last_analysis_date')),
                'reputation': attrs.get('reputation', 0),
                'tags': attrs.get('tags', []),
                'times_submitted': attrs.get('times_submitted', 0),
                'copyright': copyright_info,
                'description': description,
                'original_name': original_name,
                'product_name': product_name,
            })

            positives = malicious + suspicious
            if positives == 0:
                result['rating'] = 'clean'
            elif positives <= 5:
                result['rating'] = 'low'
            elif positives <= 15:
                result['rating'] = 'medium'
            else:
                result['rating'] = 'high'

            detections = {}
            for vendor, scan in attrs.get('last_analysis_results', {}).items():
                if scan.get('result'):
                    detections[vendor] = scan['result']
            result['detections'] = detections
            result['top_detections'] = dict(list(detections.items())[:10])

            pe_info = attrs.get('pe_info', {})
            result['imphash'] = pe_info.get('imphash', '-')

            sig_info = attrs.get('signature_info', {})
            result['signer'] = sig_info.get('signers', '-')
            result['signed'] = sig_info.get('verified', '-')

        except Exception as e:
            logger.error("Error parsing file result: %s", e)
        return result

    def _parse_ip_result(self, data, ip):
        result = {'found': True, 'type': 'ip', 'indicator': ip}
        try:
            attrs = data.get('data', {}).get('attributes', {})
            stats = attrs.get('last_analysis_stats', {})

            malicious = stats.get('malicious', 0)
            suspicious = stats.get('suspicious', 0)
            undetected = stats.get('undetected', 0)
            harmless = stats.get('harmless', 0)
            total = malicious + suspicious + undetected + harmless

            # Accurate detection display for IPs
            parts = []
            if malicious:
                parts.append(f'{malicious} malicious')
            if suspicious:
                parts.append(f'{suspicious} suspicious')
            if not parts:
                detection_display = 'Clean'
            else:
                detection_display = ', '.join(parts)

            result.update({
                'positives': malicious + suspicious,
                'total': total,
                'malicious': malicious,
                'suspicious': suspicious,
                'harmless_count': harmless,
                'detection_ratio': f'{malicious + suspicious}/{total}' if total else '0/0',
                'detection_display': detection_display,
                'country': attrs.get('country', '-'),
                'continent': attrs.get('continent', '-'),
                'asn': attrs.get('asn', '-'),
                'as_owner': attrs.get('as_owner', '-'),
                'network': attrs.get('network', '-'),
                'reputation': attrs.get('reputation', 0),
                'whois': (attrs.get('whois', '-') or '-')[:500],
                'tags': attrs.get('tags', []),
            })

            positives = malicious + suspicious
            if positives == 0:
                result['rating'] = 'clean'
            elif positives <= 3:
                result['rating'] = 'low'
            elif positives <= 10:
                result['rating'] = 'medium'
            else:
                result['rating'] = 'high'

            detections = {}
            for vendor, scan in attrs.get('last_analysis_results', {}).items():
                if scan.get('result') and scan['result'] not in ('clean', 'unrated'):
                    detections[vendor] = scan['result']
            result['detections'] = detections

        except Exception as e:
            logger.error("Error parsing IP result: %s", e)
        return result

    def _parse_domain_result(self, data, domain):
        result = {'found': True, 'type': 'domain', 'indicator': domain}
        try:
            attrs = data.get('data', {}).get('attributes', {})
            stats = attrs.get('last_analysis_stats', {})

            malicious = stats.get('malicious', 0)
            suspicious = stats.get('suspicious', 0)
            undetected = stats.get('undetected', 0)
            harmless = stats.get('harmless', 0)
            total = malicious + suspicious + undetected + harmless

            parts = []
            if malicious:
                parts.append(f'{malicious} malicious')
            if suspicious:
                parts.append(f'{suspicious} suspicious')
            if not parts:
                detection_display = 'Clean'
            else:
                detection_display = ', '.join(parts)

            result.update({
                'positives': malicious + suspicious,
                'total': total,
                'malicious': malicious,
                'suspicious': suspicious,
                'harmless_count': harmless,
                'detection_ratio': f'{malicious + suspicious}/{total}' if total else '0/0',
                'detection_display': detection_display,
                'registrar': attrs.get('registrar', '-'),
                'creation_date': self._ts(attrs.get('creation_date'), fmt='%Y-%m-%d'),
                'last_modification_date': self._ts(
                    attrs.get('last_modification_date'), fmt='%Y-%m-%d'
                ),
                'reputation': attrs.get('reputation', 0),
                'categories': attrs.get('categories', {}),
                'whois': (attrs.get('whois', '-') or '-')[:500],
                'tags': attrs.get('tags', []),
                'last_dns_records': attrs.get('last_dns_records', []),
            })

            positives = malicious + suspicious
            if positives == 0:
                result['rating'] = 'clean'
            elif positives <= 3:
                result['rating'] = 'low'
            elif positives <= 10:
                result['rating'] = 'medium'
            else:
                result['rating'] = 'high'

            detections = {}
            for vendor, scan in attrs.get('last_analysis_results', {}).items():
                if scan.get('result') and scan['result'] not in ('clean', 'unrated'):
                    detections[vendor] = scan['result']
            result['detections'] = detections

        except Exception as e:
            logger.error("Error parsing domain result: %s", e)
        return result

    @staticmethod
    def _ts(epoch, fmt='%Y-%m-%d %H:%M:%S'):
        if not epoch:
            return '-'
        try:
            return datetime.utcfromtimestamp(epoch).strftime(fmt)
        except (OSError, ValueError):
            return '-'

    # -- XLSX Generation (Font 18, Center, Middle) --

    def generate_xlsx(self, scan_id):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        job = self.scan_jobs.get(scan_id)
        if not job:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Results"

        hdr_font = Font(bold=True, color="FFFFFF", size=14)
        hdr_fill = PatternFill(start_color="2D333B", end_color="2D333B", fill_type="solid")
        data_font = Font(size=18)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        rating_fills = {
            'clean': PatternFill(start_color="3FB950", end_color="3FB950", fill_type="solid"),
            'low': PatternFill(start_color="D29922", end_color="D29922", fill_type="solid"),
            'medium': PatternFill(start_color="DB6D28", end_color="DB6D28", fill_type="solid"),
            'high': PatternFill(start_color="F85149", end_color="F85149", fill_type="solid"),
            'unknown': PatternFill(start_color="8B949E", end_color="8B949E", fill_type="solid"),
        }
        border = Border(
            left=Side(style='thin', color='30363D'),
            right=Side(style='thin', color='30363D'),
            top=Side(style='thin', color='30363D'),
            bottom=Side(style='thin', color='30363D'),
        )

        headers = [
            'Indicator', 'Type', 'Rating',
            'Malicious', 'Suspicious', 'Detection Display',
            'Comment', 'MD5', 'SHA1', 'SHA256',
            'File Type', 'File Size', 'File Names',
            'First Submission', 'Last Submission',
            'Country', 'ASN', 'AS Owner', 'Network', 'Registrar',
            'Reputation', 'Tags', 'Copyright', 'Description',
            'Signer', 'Top Detections', 'Cached', 'VT Link',
        ]

        ws.row_dimensions[1].height = 35
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center_align
            cell.border = border

        row = 2
        for indicator in job.indicators:
            r = indicator.get('result') or {}
            found = r.get('found', False)
            rating = r.get('rating', 'unknown') if found else (
                'not found' if indicator['status'] == 'completed'
                else indicator['status']
            )
            itype = indicator['type']
            vt_section = 'file' if itype in ('md5', 'sha1', 'sha256') else itype

            dets = r.get('top_detections', r.get('detections', {}))
            det_str = (
                '; '.join(f'{k}: {v}' for k, v in list(dets.items())[:10])
                if isinstance(dets, dict) else '-'
            )
            tags = r.get('tags', [])
            tag_str = ', '.join(tags) if isinstance(tags, list) else str(tags)

            values = [
                indicator['value'],
                indicator['type'].upper(),
                rating,
                r.get('malicious', 0) if found else '-',
                r.get('suspicious', 0) if found else '-',
                r.get('detection_display', '-') if found else '-',
                indicator.get('comment', ''),
                r.get('md5', '-'),
                r.get('sha1', '-'),
                r.get('sha256', '-'),
                r.get('file_type', '-'),
                r.get('file_size', '-'),
                r.get('file_names', r.get('meaningful_name', '-')),
                r.get('first_submission', '-'),
                r.get('last_submission', '-'),
                r.get('country', '-'),
                r.get('asn', '-'),
                r.get('as_owner', '-'),
                r.get('network', '-'),
                r.get('registrar', '-'),
                r.get('reputation', '-'),
                tag_str,
                r.get('copyright', '-'),
                r.get('description', '-'),
                r.get('signer', '-'),
                det_str,
                'Yes' if indicator.get('from_cache') else 'No',
                f'https://www.virustotal.com/gui/{vt_section}/{indicator["value"]}',
            ]

            ws.row_dimensions[row].height = 40
            for col, val in enumerate(values, 1):
                cell = ws.cell(
                    row=row, column=col,
                    value=str(val) if val is not None else '-',
                )
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border
                if col == 3 and rating in rating_fills:
                    cell.fill = rating_fills[rating]
                    cell.font = Font(bold=True, color="FFFFFF", size=18)

            row += 1

        widths = {
            'A': 48, 'B': 14, 'C': 16, 'D': 16, 'E': 16, 'F': 30,
            'G': 20, 'H': 36, 'I': 44, 'J': 68,
            'K': 20, 'L': 16, 'M': 35,
            'N': 24, 'O': 24,
            'P': 12, 'Q': 12, 'R': 28, 'S': 22, 'T': 28,
            'U': 14, 'V': 30, 'W': 30, 'X': 30,
            'Y': 35, 'Z': 55, 'AA': 10, 'AB': 65,
        }
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w

        ws.freeze_panes = 'A2'

        safe_name = re.sub(r'[^\w\-. ]', '_', job.name)
        xlsx_path = os.path.join(job.scan_dir, f'{safe_name}.xlsx')
        wb.save(xlsx_path)
        return xlsx_path
